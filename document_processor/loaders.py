import os
import csv

from typing import Callable
from openai import OpenAIError
from utils.llm_client import ask_audio, ask_vision
from pypdf import PdfReader

LOADER_REGISTRY: dict[str, Callable[[str], str]] = {}


def register_loader(*extensions: str):
    """Register a loader function under one or more file extensions.

    A loader is any function shaped (file_path: str) -> str — it takes a
    path and returns the raw extracted text. Adding support for a new file
    format later (Ch9: tables, images, audio, video, etc.) means writing one
    new function decorated with @register_loader(...) — this dispatch logic
    itself never needs to change.
    """
    def decorator(func: Callable[[str], str]) -> Callable[[str], str]:
        for ext in extensions:
            LOADER_REGISTRY[ext.lower()] = func
        return func
    return decorator


@register_loader(".pdf")
def load_pdf(file_path: str) -> str:
    reader = PdfReader(file_path)
    return "\n".join(page.extract_text() or "" for page in reader.pages)


@register_loader(".txt")
def load_text(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()


@register_loader(".csv")
def load_csv(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    return "\n".join(", ".join(cell for cell in row) for row in rows)


# .xlsx doesn't fit LOADER_REGISTRY's (file_path) -> str contract: a
# workbook can hold several sheets, each with its own header row, so there's
# no single string to hand back. SHEET_LOADER_REGISTRY is a separate,
# honestly-different contract for that case: (file_path) -> list of
# (sheet_name, header_line, body_text), one entry per sheet.
SHEET_LOADER_REGISTRY: dict[str, Callable[[str], list[tuple[str, str, str]]]] = {}


def register_sheet_loader(*extensions: str):
    def decorator(func):
        for ext in extensions:
            SHEET_LOADER_REGISTRY[ext.lower()] = func
        return func
    return decorator


@register_sheet_loader(".xlsx")
def load_xlsx_sheets(file_path: str) -> list[tuple[str, str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheets = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header_line = ", ".join("" if cell is None else str(cell) for cell in rows[0])
        body_lines = [
            ", ".join("" if cell is None else str(cell) for cell in row)
            for row in rows[1:]
        ]
        sheets.append((sheet.title, header_line, "\n".join(body_lines)))
    return sheets


@register_loader(".png", ".jpg", ".jpeg", ".webp")
def load_image(file_path: str) -> str:
    try:
        return ask_vision(file_path)
    except OpenAIError as e:
        raise RuntimeError(f"Image captioning failed for {file_path}: {e}") from e



@register_loader(".mp3", ".wav", ".m4a")
def load_audio(file_path: str) -> str:
    try:
        return ask_audio(file_path)
    except OpenAIError as e:
        raise RuntimeError(f"Audio transcription failed for {file_path}: {e}") from e